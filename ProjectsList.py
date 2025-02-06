from datetime import datetime
from PyQt6.QtCore import QDate
from PyQt6 import QtWidgets
from PyQt6.QtSql import QSqlDatabase, QSqlTableModel
from PyQt6.QtCore import pyqtSignal, Qt
from PyQt6.QtGui import QPixmap, QTransform
import os
import sys
from pathlib import Path
import pandas as pd
from PyQt6 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

class CenterAlignDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.displayAlignment = QtCore.Qt.AlignmentFlag.AlignCenter

class CurrencyDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.displayAlignment = QtCore.Qt.AlignmentFlag.AlignCenter

    def displayText(self, value, locale):
        if value is None:
            return ""
        try:
            # Ensure the value is a float
            value = float(value)
        except ValueError:
            return str(value)

        # Format the number with commas and two decimal places
        formatted_value = f"{value:,.2f}"
        return f"Php.{formatted_value}"

class Ui_Project_Window(QtWidgets.QWidget):
    row_data_signal = pyqtSignal(dict)
    def __init__(self):
        super().__init__()
        self.sort_orders = {}  # Dictionary to track sort orders for each column

    def setupUi(self, Project_Window):
        Project_Window.setObjectName("Project_Window")
        Project_Window.resize(1110, 590)
        Project_Window.setMinimumSize(1110, 590)
        self.centralwidget = QtWidgets.QWidget(parent=Project_Window)
        self.centralwidget.setObjectName("centralwidget")

        # Set the background image using a style sheet
        bg_image_path = Ui_Project_Window.resource_path("BGP2.png")
        Project_Window.setStyleSheet(f"""
               QMainWindow {{
                   background-image: url("{bg_image_path.replace('\\', '/')}");
                   background-repeat: no-repeat;
                   background-position: center;
                   background-size: cover;
               }}
           """)
        self.WM_lb = QtWidgets.QLabel(parent=self.centralwidget)
        self.WM_lb.setGeometry(QtCore.QRect(3, 2, 200, 16))
        self.WM_lb.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
        self.WM_lb.setObjectName("WM_lb")

        # Create main vertical layout
        self.main_layout = QtWidgets.QVBoxLayout(self.centralwidget)

        # Create a horizontal layout for the top widgets
        self.top_layout = QtWidgets.QHBoxLayout()

        self.Filter_lb = QtWidgets.QLabel(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.Filter_lb.setFont(font)
        self.Filter_lb.setObjectName("Filter_lb")
        self.Filter_lb.setFixedSize(81, 21)  # Set fixed size based on original geometry

        self.Filter_cb = QtWidgets.QComboBox(parent=self.centralwidget)
        self.Filter_cb.setMaxVisibleItems(11)
        self.Filter_cb.setMaxCount(11)
        self.Filter_cb.setObjectName("Filter_cb")
        self.Filter_cb.setFixedSize(190, 21)  # Set fixed size based on original geometry
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")
        self.Filter_cb.addItem("")

        self.Search_lb = QtWidgets.QLabel(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.Search_lb.setFont(font)
        self.Search_lb.setObjectName("Search_lb")
        self.Search_lb.setFixedSize(51, 21)  # Set fixed size based on original geometry

        self.Search_Line = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.Search_Line.setObjectName("Search_Line")
        self.Search_Line.setFixedSize(190, 21)  # Set fixed size based on original geometry

        self.Search_btn = QtWidgets.QPushButton(parent=self.centralwidget)
        self.Search_btn.setObjectName("Search_btn")
        self.Search_btn.setFixedSize(75, 24)  # Set fixed size based on original geometry
        self.Search_btn.clicked.connect(self.apply_filter)
        self.Search_Line.returnPressed.connect(self.apply_filter)

        self.Show_btn = QtWidgets.QPushButton(parent=self.centralwidget)
        self.Show_btn.setObjectName("Show_btn")
        self.Show_btn.setFixedSize(131, 52)  # Set fixed size based on original geometry

        self.Revert_btn = QtWidgets.QPushButton(parent=self.centralwidget)
        self.Revert_btn.setObjectName("Revert_btn")
        self.Revert_btn.setFixedSize(131, 52)  # Set fixed size based on original geometry
        self.Revert_btn.setVisible(False)  # Make Revert_btn invisible initially

        self.Transfer_btn = QtWidgets.QPushButton(parent=self.centralwidget)
        self.Transfer_btn.setObjectName("Transfer_btn")
        self.Transfer_btn.setFixedSize(131, 52)  # Set fixed size based on original geometry
        self.Transfer_btn.setVisible(False)  # Make Transfer_btn invisible initially

        self.Clear_btn = QtWidgets.QPushButton(parent=self.centralwidget)
        self.Clear_btn.setObjectName("Clear_btn")
        self.Clear_btn.setFixedSize(75, 24)  # Set fixed size based on original geometry
        self.Refresh_btn = QtWidgets.QPushButton(parent=self.centralwidget)
        self.Refresh_btn.setObjectName("Refresh_btn")
        self.Refresh_btn.setFixedSize(75, 24)  # Set fixed size based on original geometry

        # Add widgets to the top layout with spacers
        self.top_layout.addWidget(self.Filter_lb)
        self.top_layout.addItem(
            QtWidgets.QSpacerItem(10, 0, QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Minimum))
        self.top_layout.addWidget(self.Filter_cb)
        self.top_layout.addItem(
            QtWidgets.QSpacerItem(10, 0, QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Minimum))
        self.top_layout.addWidget(self.Search_lb)
        self.top_layout.addItem(
            QtWidgets.QSpacerItem(10, 0, QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Minimum))
        self.top_layout.addWidget(self.Search_Line)
        self.top_layout.addItem(
            QtWidgets.QSpacerItem(10, 0, QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Minimum))
        self.top_layout.addWidget(self.Search_btn)
        self.top_layout.addItem(
            QtWidgets.QSpacerItem(2, 0, QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Minimum))  # Shorten the gap to 5px
        self.top_layout.addWidget(self.Clear_btn)
        self.top_layout.addItem(
            QtWidgets.QSpacerItem(2, 0, QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Minimum))
        self.top_layout.addWidget(self.Refresh_btn)
        self.top_layout.addItem(
            QtWidgets.QSpacerItem(600, 0, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum))
        self.top_layout.addWidget(self.Revert_btn)  # Revert_btn is now added before Show_btn
        self.top_layout.addItem(
            QtWidgets.QSpacerItem(10, 0, QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Minimum))
        self.top_layout.addWidget(self.Show_btn)  # Show_btn is now added after Revert_btn
        self.top_layout.addWidget(self.Transfer_btn)  # Add Transfer_btn to the layout

        # Add top layout to the main layout
        self.main_layout.addLayout(self.top_layout)

        # Add DBViewer to the main layout and set it to expand
        self.DBViewer = QtWidgets.QTableView(parent=self.centralwidget)
        self.DBViewer.setObjectName("DBViewer")
        self.DBViewer.verticalHeader().setVisible(False)  # Hide the row numbering column
        self.DBViewer.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.DBViewer.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.DBViewer.doubleClicked.connect(self.handle_row_double_click)
        self.main_layout.addWidget(self.DBViewer)

        Project_Window.setCentralWidget(self.centralwidget)

        self.retranslateUi(Project_Window)
        self.Filter_cb.setCurrentIndex(-1)
        QtCore.QMetaObject.connectSlotsByName(Project_Window)
        self.DBViewer.setSelectionBehavior(QtWidgets.QTableView.SelectionBehavior.SelectRows)
        self.Refresh_btn.clicked.connect(self.refresh_data)

        self.setupDatabase()
        self.setupModel()
        self.Show_btn.clicked.connect(lambda: QtCore.QTimer.singleShot(800, self.show_transfer_revert))
        self.Show_btn.clicked.connect(lambda: QtCore.QTimer.singleShot(800, self.show_transfer_revert))
        self.Revert_btn.clicked.connect(lambda: QtCore.QTimer.singleShot(800, self.revert_show_transfer))
        self.Transfer_btn.clicked.connect(lambda: QtCore.QTimer.singleShot(800, self.transfer_data_to_excel))

        self.DBViewer.horizontalHeader().sectionClicked.connect(self.sort_columns)
        self.Filter_cb.currentIndexChanged.connect(self.apply_filter)
        self.Clear_btn.clicked.connect(self.clear_filter)

    def retranslateUi(self, Project_Window):
        _translate = QtCore.QCoreApplication.translate
        Project_Window.setWindowTitle(_translate("Project_Window", "Projects List"))
        self.WM_lb.setText(_translate("Project_Window", "Developed by: J. Masunsong"))
        self.Filter_lb.setText(_translate("Project_Window", "Filtered by:"))
        self.Filter_cb.setItemText(0, _translate("Project_Window", "Tracking Number"))
        self.Filter_cb.setItemText(1, _translate("Project_Window", "Project Year"))
        self.Filter_cb.setItemText(2, _translate("Project_Window", "Project/Activity"))
        self.Filter_cb.setItemText(3, _translate("Project_Window", "Location"))
        self.Filter_cb.setItemText(4, _translate("Project_Window", "Notice to Proceed"))
        self.Filter_cb.setItemText(5, _translate("Project_Window", "Date Started"))
        self.Filter_cb.setItemText(6, _translate("Project_Window", "Target Completion Date"))
        self.Filter_cb.setItemText(7, _translate("Project_Window", "Project Status (%)"))
        self.Filter_cb.setItemText(8, _translate("Project_Window", "Inspection Date"))
        self.Filter_cb.setItemText(9, _translate("Project_Window", "Project Coordinator"))
        self.Filter_cb.setItemText(10, _translate("Project_Window", "Source of Fund"))
        self.Filter_cb.setItemText(11, _translate("Project_Window", "Contractor"))
        self.Search_lb.setText(_translate("Project_Window", "Search:"))
        self.Search_btn.setText(_translate("Project_Window", "Search"))
        self.Refresh_btn.setText(_translate("Project_Window", "Refresh"))
        self.Clear_btn.setText(_translate("Project_Window", "Clear"))
        self.Show_btn.setText(_translate("Project_Window", "Show Full Table"))
        self.Revert_btn.setText(_translate("Project_Window", "Revert"))
        self.Transfer_btn.setText(_translate("Project_Window", "Transfer Data to Excel"))

    @staticmethod
    def resource_path(relative_path):
        """Get the absolute path to a resource, works for PyInstaller."""
        try:
            # For when the application is bundled with PyInstaller
            base_path = sys._MEIPASS
        except AttributeError:
            # For regular Python execution
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def setupDatabase(self):
        self.db = QSqlDatabase.addDatabase('QMYSQL')
        self.db.setHostName('192.168.141.214')
        self.db.setDatabaseName('dmedb')
        self.db.setUserName('appuser')
        self.db.setPassword('StrongPassword123!')
        if not self.db.open():
            QtWidgets.QMessageBox.critical(None, "Database Error", self.db.lastError().text())

    def setupModel(self):
        self.model = QSqlTableModel()
        self.model.setTable('project_tb')
        self.model.setEditStrategy(QSqlTableModel.EditStrategy.OnFieldChange)
        self.model.select()
        self.DBViewer.setModel(self.model)
        self.DBViewer.setModel(self.model)
        self.deleted_row_ids = set()
        # Hide specific columns by name
        self.columns_to_hide = [
            "project_year", "date_extension",
            "project_photos", "project_remarks", "date_days"
        ]
        self.permanently_hidden_columns = ["project_id"]

        for column in self.columns_to_hide + self.permanently_hidden_columns:
            index = self.model.fieldIndex(column)
            if index != -1:
                self.DBViewer.setColumnHidden(index, True)

        # Set custom headers for specific columns
        headers = {
            "tracking_number": "Tracking Number",
            "project_year": "Project Year",
            "project_title": "Project/Activity Title",
            "project_location": "Location",
            "project_totalcost": "Bid Cost",
            "project_budget": "Approved Budget of Contract (ABC)",
            "date_notice": "Notice of Award",
            "date_start": "Date Started",
            "date_days": "Duration of Calendar Days",
            "date_extension": "No. of Extension",
            "date_target": "Target Completion Date",
            "project_status": "Project Status (%)",
            "project_incurred": "Total Cost Incurred to Date",
            "date_inspection": "Latest Inspection Date",
            "project_photos": "Photos",
            "project_remarks": "Remarks",
            "project_coordinator": "Project Coordinator",
            "project_source": "Source of Fund",
            "project_contractor": "Contractor",
            "project_encoder": "Encoder"
        }

        for column, header in headers.items():
            index = self.model.fieldIndex(column)
            if index != -1:
                self.model.setHeaderData(index, QtCore.Qt.Orientation.Horizontal, header)

        self.adjustColumnWidths()

        # Set the center alignment delegate for all columns
        center_delegate = CenterAlignDelegate(self.DBViewer)
        self.DBViewer.setItemDelegate(center_delegate)

        # Set the currency delegate for specific columns
        currency_delegate = CurrencyDelegate(self.DBViewer)
        self.DBViewer.setItemDelegateForColumn(self.model.fieldIndex("project_totalcost"), currency_delegate)
        self.DBViewer.setItemDelegateForColumn(self.model.fieldIndex("project_budget"), currency_delegate)
        self.DBViewer.setItemDelegateForColumn(self.model.fieldIndex("project_incurred"), currency_delegate)

    def adjustColumnWidths(self):
        # Adjust column widths based on the header text length and cell content length
        font_metrics = self.DBViewer.fontMetrics()
        columns_to_adjust_by_content = [
            "project_title", "project_location", "project_totalcost",
            "project_coordinator", "project_source", "project_contractor",
            "project_encoder"
        ]
        columns_to_adjust_by_header = [
            "tracking_number", "project_budget", "date_notice", "date_days", "date_target", "project_incurred"
        ]

        for column_name in columns_to_adjust_by_content:
            index = self.model.fieldIndex(column_name)
            if index != -1:
                max_width = 0
                for row in range(self.model.rowCount()):
                    text = self.model.data(self.model.index(row, index))
                    if text:
                        text = str(text)  # Ensure the text is a string
                        max_width = max(max_width, font_metrics.horizontalAdvance(text))
                max_width += 20  # Add some padding
                self.DBViewer.setColumnWidth(index, max_width)

        for column_name in columns_to_adjust_by_header:
            index = self.model.fieldIndex(column_name)
            if index != -1:
                header_text = self.model.headerData(index, QtCore.Qt.Orientation.Horizontal)
                max_width = font_metrics.horizontalAdvance(header_text) + 20  # Add some padding
                self.DBViewer.setColumnWidth(index, max_width)

        total_cost_index = self.model.fieldIndex("project_totalcost")
        if total_cost_index != -1:
            test_value = "Php. 123,456,789.00"
            required_width = font_metrics.horizontalAdvance(test_value) + 20  # Add some padding
            self.DBViewer.setColumnWidth(total_cost_index, required_width)

    def show_transfer_revert(self):
        self.Revert_btn.setVisible(True)
        self.Show_btn.setVisible(False)
        self.Transfer_btn.setVisible(True)

        # Hide the "Encoder" column (project_encoder)
        encoder_index = self.model.fieldIndex("project_encoder")
        if encoder_index != -1:
            self.DBViewer.setColumnHidden(encoder_index, True)

        # Show all other hidden columns except permanently hidden ones
        for column in self.columns_to_hide:
            index = self.model.fieldIndex(column)
            if index != -1:
                self.DBViewer.setColumnHidden(index, False)

    def revert_show_transfer(self):
        self.Revert_btn.setVisible(False)
        self.Show_btn.setVisible(True)
        self.Transfer_btn.setVisible(False)

        # Show the "Encoder" column (project_encoder) again
        encoder_index = self.model.fieldIndex("project_encoder")
        if encoder_index != -1:
            self.DBViewer.setColumnHidden(encoder_index, False)

        # Re-hide the other columns in the list
        for column in self.columns_to_hide:
            index = self.model.fieldIndex(column)
            if index != -1:
                self.DBViewer.setColumnHidden(index, True)

    def transfer_data_to_excel(self):
        # Define fixed column widths (in pixels, A to R)
        column_widths = {
            "A": 100, "B": 55, "C": 105, "D": 100, "E": 120, "F": 120,
            "G": 80, "H": 80, "I": 65, "J": 70, "K": 80, "L": 60,
            "M": 120, "N": 80, "O": 50, "P": 80, "Q": 90, "R": 90, "S": 90,
        }

        # Columns to exclude from the Excel file
        excluded_headers = ["project_id" ,"Encoder"]  # Exclude Encoder from the exported file

        # Get database table model, header names, and data
        model = self.model
        headers = []
        data = []

        # Map visible columns (excluding hidden ones)
        visible_columns = []
        for col in range(model.columnCount()):
            header = model.headerData(col, QtCore.Qt.Orientation.Horizontal)
            if header not in excluded_headers:  # Skip excluded fields
                headers.append(header)
                visible_columns.append(col)  # Keep track of visible column indexes

        # Extract data only for visible columns
        for row in range(model.rowCount()):
            row_data = []
            for col in visible_columns:  # Only iterate through visible columns
                value = model.index(row, col).data()
                if isinstance(value, QDate):
                    value = value.toString("yyyy-MM-dd")
                row_data.append(value)
            data.append(row_data)

        # Load the template workbook
        downloads_dir = str(Path.home() / "Downloads")
        workbook = Workbook()
        workbook.create_sheet("Sheet1", index=0)
        sheet = workbook.active  # Active sheet is modified

        #Define cell style
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Write data rows
        for row_num, row_data in enumerate(data, start=2):
            for col_num, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = center_alignment
                cell.border = thin_border

                # Adjust formatting based on column names (headers)
                header = headers[col_num - 1]

                # Specific formatting for "Total Cost Incurred to Date"
                if header == "Total Cost Incurred to Date":
                    cell.number_format = '"Php"#,##0.00'

                # General formatting for columns containing "Cost" or "Budget"
                elif "Cost" in header or "Budget" in header:
                    cell.number_format = '"Php"#,##0.00'

                # General formatting for date columns (MM/DD/YYYY format)
                elif "Date" in header or "Proceed" in header:
                    if isinstance(value, str) and "-" in value:  # Assuming date format is YYYY-MM-DD
                        try:
                            # Convert to datetime object then to MM/DD/YYYY format
                            value = datetime.strptime(value, "%Y-%m-%d").strftime("%m/%d/%Y")
                        except ValueError:
                            pass  # Keep the value as-is if it cannot be formatted
                    cell.value = value
                    cell.number_format = "MM/DD/YYYY"  # Apply Excel number format

        # Set column widths from A to R
        for col_letter, width in column_widths.items():
            # Divide the pixel width by 7.5 (Excel column width unit conversion)
            sheet.column_dimensions[col_letter].width = width / 7.5

            # Add 2 blank rows at the end of data
            last_row = len(data) + 1
            sheet.append([])
            sheet.append([])

            # Add "Prepared by:" in the 3rd row below the last row of data
            prepared_by_row = last_row + 3
            sheet.cell(row=prepared_by_row, column=3).value = "Prepared by:"
            sheet.cell(row=prepared_by_row, column=3).alignment = center_alignment
            sheet.cell(row=prepared_by_row, column=3).font = Font(bold=True)

            approval_by_row = last_row + 3
            sheet.cell(row=approval_by_row, column=5).value = "Noted by:"
            sheet.cell(row=approval_by_row, column=5).alignment = center_alignment
            sheet.cell(row=approval_by_row, column=5).font = Font(bold=True)

            merged_row_start = prepared_by_row + 2
            mergedapproval_row = merged_row_start

            # Merge cells C and D for the first merged row and add "<Insert Name>"
            sheet.merge_cells(start_row=merged_row_start, start_column=3, end_row=merged_row_start, end_column=4)
            cell_name = sheet.cell(row=merged_row_start, column=3)
            cell_name.value = "<Insert Name>"
            cell_name.alignment = center_alignment
            cell_name.font = Font(bold=True)

            # Merge cells E and F for the first merged row
            sheet.merge_cells(start_row=mergedapproval_row, start_column=5, end_row=merged_row_start, end_column=6)
            cell_name = sheet.cell(row=mergedapproval_row, column=5)
            cell_name.value = "Engr. Voltaire M. Villela"
            cell_name.alignment = center_alignment
            cell_name.font = Font(bold=True)

            # Merge cells C and D for the second merged row and add "<Insert Position>"
            sheet.merge_cells(start_row=merged_row_start + 1, start_column=3, end_row=merged_row_start + 1, end_column=4)
            cell_position = sheet.cell(row=merged_row_start + 1, column=3)
            cell_position.value = "<Insert Position>"
            cell_position.alignment = center_alignment
            cell_position.font = Font(italic=True)

            # Merge cells E and F for the first merged row
            sheet.merge_cells(start_row=mergedapproval_row + 1, start_column=5, end_row=merged_row_start + 1, end_column=6)
            cell_name = sheet.cell(row=mergedapproval_row + 1, column=5)
            cell_name.value = "Municipal Engineer, OIC"
            cell_name.alignment = center_alignment
            cell_name.font = Font(italic=True)

        # Save the updated workbook to a new file
        updated_file_path = os.path.join(downloads_dir, "Table List.xlsx")
        workbook.save(updated_file_path)
        QtWidgets.QMessageBox.information(
            None, "Success", "Data has been successfully exported to 'Table List.xlsx' in Downloads."
        )

    def sort_columns(self, logicalIndex):
        # Toggle the sort order for the clicked column
        if logicalIndex in self.sort_orders:
            self.sort_orders[logicalIndex] = not self.sort_orders[logicalIndex]
        else:
            self.sort_orders[logicalIndex] = True  # Default to ascending order

        order = QtCore.Qt.SortOrder.AscendingOrder if self.sort_orders[logicalIndex] else QtCore.Qt.SortOrder.DescendingOrder
        self.DBViewer.sortByColumn(logicalIndex, order)

    def apply_filter(self):
        # Get the selected filter
        filter_text = self.Filter_cb.currentText()

        # Get the search text
        search_text = self.Search_Line.text().strip()

        # Check if the Revert_btn is visible or not
        if filter_text == "Project Year" and not self.Revert_btn.isVisible():
            QtWidgets.QMessageBox.warning(None, "Warning",
                                          "Click 'Show Full Table' first before filtering by Project Year.")
            return

        # Check if the Revert_btn is visible and Show_btn is invisible for filtering Project Year
        if filter_text == "Project Year" and self.Revert_btn.isVisible() and not self.Show_btn.isVisible():
            # Apply filter for project_year based on Search_Line text
            search_text = self.Search_Line.text()
            self.model.setFilter(f"project_year LIKE '%{search_text}%'")

            # Apply different filters based on the length of search text
            if len(search_text) <= 2:
                # Show rows where the last two characters of the first word match the search_text
                self.model.setFilter(f"project_year LIKE '%{search_text}%'")

            self.model.select()
            return

        # Define a dictionary mapping filter texts to corresponding date fields
        date_filter_mapping = {
            "Notice to Proceed": "date_notice",
            "Date Started": "date_start",
            "Target Completion Date": "date_target",
            "Inspection Date": "date_inspection"
        }

        if filter_text in date_filter_mapping:
            column_name = date_filter_mapping[filter_text]

            # Handle numerical month input by padding single digits with a leading zero
            if search_text.isdigit() and 1 <= int(search_text) <= 12:
                search_text = search_text.zfill(2)

            # Use DATE_FORMAT to extract the month name, abbreviated month name, and day from the date field
            if ' ' in search_text:
                try:
                    month, day = search_text.split(' ')
                    day = day.zfill(2)
                    self.model.setFilter(
                        f"(DATE_FORMAT({column_name}, '%M %d') = '{month} {day}' "
                        f"OR DATE_FORMAT({column_name}, '%b %d') = '{month} {day}')"
                    )
                except ValueError:
                    self.model.setFilter("")
            else:
                if search_text:
                    self.model.setFilter(
                        f"DATE_FORMAT({column_name}, '%M') = '{search_text}' "
                        f"OR DATE_FORMAT({column_name}, '%b') = '{search_text}' "
                        f"OR DATE_FORMAT({column_name}, '%m') = '{search_text}'"
                    )
                else:
                    self.model.setFilter("")
        else:
            # Determine the filter condition based on the length of search text
            if len(search_text) <= 2:
                filter_condition = f"{search_text}%"
            else:
                filter_condition = f"%{search_text}%"

            # Apply the filter for non-date fields
            if filter_text == "Project Year":
                self.model.setFilter(f"project_year LIKE '{filter_condition}'")
            elif filter_text == "Tracking Number":
                self.model.setFilter(f"tracking_number LIKE '{filter_condition}'")
            elif filter_text == "Project/Activity":
                self.model.setFilter(f"project_title LIKE '{filter_condition}'")
            elif filter_text == "Location":
                self.model.setFilter(f"project_location LIKE '{filter_condition}'")
            elif filter_text == "Project Status (%)":
                self.model.setFilter(f"project_status LIKE '{filter_condition}'")
            elif filter_text == "Project Coordinator":
                self.model.setFilter(f"project_coordinator LIKE '{filter_condition}'")
            elif filter_text == "Source of Fund":
                self.model.setFilter(f"project_source LIKE '{filter_condition}'")
            elif filter_text == "Contractor":
                self.model.setFilter(f"project_contractor LIKE '{filter_condition}'")
            else:
                self.model.setFilter("")

        self.model.select()

    def refresh_data(self):
        self.model.select()

    def clear_filter(self):
        self.model.setFilter("")
        self.model.select()
        self.Filter_cb.setCurrentIndex(-1)
        self.Search_Line.clear()
        self.Search_Line.clear()

    def handle_row_double_click(self, index):
        # Get the row number of the double-clicked cell
        row = index.row()
        row_data = {}

        # Loop through all columns of the clicked row
        for column in range(self.model.columnCount()):
            column_name = self.model.headerData(column, QtCore.Qt.Orientation.Horizontal)
            cell_data = self.model.index(row, column).data()
            row_data[column_name] = cell_data

        print(f"Emitting row data: {row_data}") ######### DEBUGGER #########

        self.row_data_signal.emit(row_data)

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    Project_Window = QtWidgets.QMainWindow()
    ui = Ui_Project_Window()
    ui.setupUi(Project_Window)
    Project_Window.show()
    sys.exit(app.exec())